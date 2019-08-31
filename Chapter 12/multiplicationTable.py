import openpyxl, sys, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

n = int(sys.argv[1])
cwd = os.chdir(input('Save to Directory: '))
wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, n + 1):
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=1, column=i + 1).font = Font(bold=True)
    sheet.cell(row=i + 1, column=1).font = Font(bold=True)

for eachRow in range(1, n + 1):
    for eachCol in range(2, int(sheet.max_column) + 1):
        n1 = sheet['A' + str(eachRow + 1)].value
        n2 = sheet[get_column_letter(eachCol) + '1'].value
        sheet[get_column_letter(eachCol) + str(eachRow + 1)].value = n1 * n2

wb.save('multiplicationTable.xlsx')
print('Multiplication Table is complete.')
