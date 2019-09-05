import openpyxl

wb = openpyxl.load_workbook('invertersample.xlsx')
sheet = wb.active
new_sheet = wb.create_sheet('invertedSheet')

# Create an inverted data table in a new sheet within the same workbook.
for x in range(1, sheet.max_row + 1):
    for y in range(1, sheet.max_column + 1):
        new_sheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value

wb.save('invertersample2.xlsx')
