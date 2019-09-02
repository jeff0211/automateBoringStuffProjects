import openpyxl, sys, os

os.chdir(input('Workbook Directory: '))

n = int(sys.argv[1])
m = int(sys.argv[2])
wb = openpyxl.load_workbook(sys.argv[3])
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Retaining the rows of data before insertion.
for i in range(0, n - 1):
    for item in sheet.columns:
        new_sheet[item[i].coordinate] = item[i].value

# Insert the remaining rows of data starting at row n+m.
for x in range(n - 1, sheet.max_row):
    colCount = 1
    for item in sheet.columns:
        new_sheet.cell(row=x + m + 1, column=colCount).value = item[x].value
        colCount += 1

# Saving file to current working directory
print('Blank rows insertion complete.')
fileName = input('Enter New File Name (Without Ext): ')
new_wb.save(fileName + '.xlsx')
print('File Saved.')
